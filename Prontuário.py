import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook
import os

# Função para salvar dados em um arquivo Excel
def salvar_em_excel(dados):
    # Se o arquivo já existir, carregar, senão criar um novo
    if os.path.exists("prontuarios.xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook("prontuarios.xlsx")
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Criar cabeçalho na primeira linha
        ws.append(["Nome", "Data de Nascimento", "Idade", "Sexo", "Endereço", "Telefone", "Queixa Principal", "Procedimento", "Data do Procedimento"])

    # Adicionar os dados na próxima linha
    ws.append(dados)

    # Salvar o arquivo
    wb.save("prontuarios.xlsx")
    wb.close()

# Função que será chamada quando o botão "Salvar" for clicado
def salvar_dados():
    nome = entry_nome.get()
    data_nascimento = entry_data_nascimento.get()
    idade = entry_idade.get()
    sexo = var_sexo.get()
    endereco = entry_endereco.get()
    telefone = entry_telefone.get()
    queixa = entry_queixa.get("1.0", tk.END).strip()
    procedimento = var_procedimento.get()
    data_procedimento = entry_data_procedimento.get()

    # Validar se todos os campos foram preenchidos
    if nome and data_nascimento and idade and sexo and endereco and telefone and queixa and procedimento and data_procedimento:
        # Dados a serem salvos no Excel
        dados = [nome, data_nascimento, idade, sexo, endereco, telefone, queixa, procedimento, data_procedimento]
        
        # Salvar os dados no Excel
        salvar_em_excel(dados)

        # Adicionar os dados na Treeview
        tree.insert("", "end", values=dados)

        # Exibir mensagem de sucesso
        messagebox.showinfo("Sucesso", "Prontuário salvo com sucesso no Excel!")
        
        # Limpar os campos do formulário
        entry_nome.delete(0, tk.END)
        entry_data_nascimento.delete(0, tk.END)
        entry_idade.delete(0, tk.END)
        entry_endereco.delete(0, tk.END)
        entry_telefone.delete(0, tk.END)
        entry_queixa.delete("1.0", tk.END)
        var_procedimento.set(None)
        entry_data_procedimento.delete(0, tk.END)
    else:
        messagebox.showwarning("Aviso", "Por favor, preencha todos os campos.")

# Configuração da janela principal
root = tk.Tk()
root.title("Preenchimento de Prontuário")
root.geometry("900x700")

# Labels e entradas para o nome
tk.Label(root, text="Nome do Paciente:").pack(pady=5)
entry_nome = tk.Entry(root, width=80)
entry_nome.pack()

# Labels e entradas para a data de nascimento
tk.Label(root, text="Data de Nascimento (DD/MM/AAAA):").pack(pady=5)
entry_data_nascimento = tk.Entry(root, width=40)
entry_data_nascimento.pack()

# Labels e entradas para a idade
tk.Label(root, text="Idade:").pack(pady=5)
entry_idade = tk.Entry(root, width=40)
entry_idade.pack()

# Labels e opções de seleção para o sexo
tk.Label(root, text="Sexo:").pack(pady=5)
var_sexo = tk.StringVar(value="Masculino")
frame_sexo = tk.Frame(root)
frame_sexo.pack(pady=5)
tk.Radiobutton(frame_sexo, text="Masculino", variable=var_sexo, value="Masculino").pack(side=tk.LEFT)
tk.Radiobutton(frame_sexo, text="Feminino", variable=var_sexo, value="Feminino").pack(side=tk.LEFT)
tk.Radiobutton(frame_sexo, text="Outro", variable=var_sexo, value="Outro").pack(side=tk.LEFT)

# Label e entrada para o endereço
tk.Label(root, text="Endereço:").pack(pady=5)
entry_endereco = tk.Entry(root, width=80)
entry_endereco.pack()

# Label e entrada para o telefone
tk.Label(root, text="Telefone:").pack(pady=5)
entry_telefone = tk.Entry(root, width=40)
entry_telefone.pack()

# Label e entrada para a queixa principal
tk.Label(root, text="Queixa Principal:").pack(pady=5)
entry_queixa = tk.Text(root, width=40, height=4)
entry_queixa.pack()

# Label para o procedimento
tk.Label(root, text="Procedimento:").pack(pady=5)
var_procedimento = tk.StringVar(value=None)

# Radiobuttons para selecionar o procedimento
frame_procedimento = tk.Frame(root)
frame_procedimento.pack(pady=5)
tk.Radiobutton(frame_procedimento, text="Tratamento de Microvasos", variable=var_procedimento, value="Tratamento de Microvasos").pack(anchor=tk.W)
tk.Radiobutton(frame_procedimento, text="Bioestimulador", variable=var_procedimento, value="Bioestimulador").pack(anchor=tk.W)
tk.Radiobutton(frame_procedimento, text="Botox", variable=var_procedimento, value="Botox").pack(anchor=tk.W)

# Label e entrada para a data do procedimento
tk.Label(root, text="Data do Procedimento (DD/MM/AAAA):").pack(pady=5)
entry_data_procedimento = tk.Entry(root, width=40)
entry_data_procedimento.pack()

# Botão para salvar os dados
btn_salvar = tk.Button(root, text="Salvar Prontuário", command=salvar_dados)
btn_salvar.pack(pady=20)

# Criação da Treeview para exibir os dados
tree_frame = tk.Frame(root)
tree_frame.pack(pady=20)

# Definir a árvore com colunas
tree = ttk.Treeview(tree_frame, columns=("Nome", "Data de Nascimento", "Idade", "Sexo", "Endereço", "Telefone", "Queixa Principal", "Procedimento", "Data do Procedimento"), show="headings")

# Definir os títulos das colunas
tree.heading("Nome", text="Nome")
tree.heading("Data de Nascimento", text="Data de Nascimento")
tree.heading("Idade", text="Idade")
tree.heading("Sexo", text="Sexo")
tree.heading("Endereço", text="Endereço")
tree.heading("Telefone", text="Telefone")
tree.heading("Queixa Principal", text="Queixa Principal")
tree.heading("Procedimento", text="Procedimento")
tree.heading("Data do Procedimento", text="Data do Procedimento")

# Definir o tamanho das colunas
tree.column("Nome", width=100)
tree.column("Data de Nascimento", width=100)
tree.column("Idade", width=50)
tree.column("Sexo", width=70)
tree.column("Endereço", width=150)
tree.column("Telefone", width=100)
tree.column("Queixa Principal", width=150)
tree.column("Procedimento", width=150)
tree.column("Data do Procedimento", width=150)

# Inserir a Treeview na janela
tree.pack()

# Iniciar a aplicação
root.mainloop()
